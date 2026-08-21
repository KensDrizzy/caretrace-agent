from __future__ import annotations

import smtplib # 发送 SMTP 邮件
import ssl # 建立安全 TLS/SSL 连接
import threading # 创建线程锁，保护 Excel 写入
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.core.config import Settings
from app.core.enums import RiskCaseStatus, ToolStatus
from app.models.entities import AlertRecord, CaseNote, ExcelRecord, PsychologicalReport, RiskCase, UserAccount, now
from app.services.skills import MindBridgeSkillLibrary


EXCEL_WRITE_LOCK = threading.Lock()

# 负责处理高风险心理报告产生后的后台动作
# 这是工具编排服务，封装了 Excel、个案、邮件和备注等后台动作。
class ToolOrchestrationService:
    def __init__(self, db: Session, settings: Settings):
        self.db = db
        self.settings = settings

# 把心理报告写入 Excel 风险台账，并在数据库中记录写入结果。
    def write_excel(self, report: PsychologicalReport) -> ExcelRecord:
        # 幂等设计，检查是否已经成功写过：查询数据库中是否已经存在：同一个 report_id且状态为 SUCCESS
        existing = (
            self.db.query(ExcelRecord)
            .filter(ExcelRecord.report_id == report.id, ExcelRecord.status == ToolStatus.SUCCESS.value)
            .first()
        )
        # 如果已经写过报告了，直接返回，不重复写 Excel。
        if existing is not None:
            return existing
        # 获取 Excel 路径
        path = Path(self.settings.excel_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        # 加锁写 Excel，保证同一个进程内同一时间只有一个线程修改 Excel。
        with EXCEL_WRITE_LOCK:
            # 打开已有excel文件或创建新文件
            if path.exists():
                workbook = load_workbook(path)
                sheet = workbook.active
            else:
                # 如果没有就创建新工作簿，并设置：工作表名：CareTrace Risk Ledger
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "CareTrace Risk Ledger"
                sheet.append(["reportId", "riskLevel", "emotion", "confidence", "summary", "createdAt"])
            # 把心理报告转成一行 Excel 数据
            sheet.append([report.id, report.risk_level, report.emotion, report.confidence, report.summary, report.created_at.isoformat()])
            # 保存 Excel
            workbook.save(path)
        # 记录数据库状态
        record = ExcelRecord(report_id=report.id, file_path=str(path), status=ToolStatus.SUCCESS.value, message="Excel 台账已写入")
        self.db.add(record)
        self.db.commit()
        return record

# 创建风险个案：根据心理报告创建一个需要人工跟进的风险个案。
    def create_case(self, report: PsychologicalReport) -> RiskCase:
        # 检查是否已经创建
        existing = self.db.query(RiskCase).filter(RiskCase.report_id == report.id).first()
        # 如果该报告已经有个案：直接返回原个案，避免重复创建
        if existing is not None:
            return existing
        # 查询学生信息
        user = self.db.get(UserAccount, report.user_id)
        # 创建个案对象
        case = RiskCase(
            report_id=report.id,
            risk_level=report.risk_level,
            status=RiskCaseStatus.OPEN.value,
            owner=self._primary_owner(),
            summary=report.summary,
            # MindBridgeSkillLibrary.counselor_handoff_summary会根据 Skill 模板生成后台交接内容
            handoff_summary=MindBridgeSkillLibrary.counselor_handoff_summary(report, user),
        )
        self.db.add(case)
        self.db.commit()
        return case

# 发送个案预警：根据风险个案找到对应报告，发送或记录预警，并更新个案状态。
    def send_case_alert(self, case: RiskCase) -> AlertRecord:
        report = self.db.get(PsychologicalReport, case.report_id)
        if report is None:
            raise RuntimeError(f"report {case.report_id} not found")
        record = self.notify(report, case)
        if record.status == ToolStatus.SUCCESS.value and case.status == RiskCaseStatus.OPEN.value:
            case.status = RiskCaseStatus.ALERT_SENT.value
        case.updated_at = now()
        self.db.add(case)
        self.db.commit()
        return record

# 确认接手个案：辅导员或管理员确认已经接手这个风险个案。
    def acknowledge_case(self, case_id: int, actor: str, note: str = "") -> RiskCase:
        # 查询个案
        case = self.db.get(RiskCase, case_id)
        if case is None:
            raise RuntimeError(f"case {case_id} not found")
        # 清洗接手人名称，如果传入空字符串，就记录为：unknown
        actor_name = actor.strip() or "unknown"
        # 状态变为 ACKNOWLEDGED
        case.status = RiskCaseStatus.ACKNOWLEDGED.value
        case.acknowledged_by = actor_name
        case.acknowledged_at = now()
        case.updated_at = now()
        self.db.add(case)
        # 生成接手备注
        self._add_case_note(case.id, actor_name, note.strip() or "已确认接手该个案")
        self.db.commit()
        return case

# 添加个案备注：add_case_note，给已有风险个案添加一条人工处理记录。
    def add_case_note(self, case_id: int, actor: str, note: str) -> CaseNote:
        # 查询个案
        case = self.db.get(RiskCase, case_id)
        if case is None:
            raise RuntimeError(f"case {case_id} not found")
        # 更新个案时间
        case.updated_at = now()
        self.db.add(case)
        # 创建备注
        record = self._add_case_note(case.id, actor.strip() or "unknown", note.strip())
        self.db.commit()
        return record

# 邮件/日志通知的核心方法
    def notify(self, report: PsychologicalReport, case: RiskCase | None = None) -> AlertRecord:
        # 如果同一个报告已经成功发送或记录过：直接返回，不重复通知。
        existing = (
            self.db.query(AlertRecord)
            .filter(AlertRecord.report_id == report.id, AlertRecord.status == ToolStatus.SUCCESS.value)
            .first()
        )
        if existing is not None:
            return existing
        # 获取收件人和投递模式
        recipient = self.settings.alert_email_to.strip() or "unconfigured"
        mode = self.settings.alert_email_delivery_mode.strip().lower()
        # log 模式不真正发邮件，只在数据库中写一条成功记录。
        if mode == "log":
            return self._save_alert(
                report,
                recipient if recipient != "unconfigured" else "log",
                ToolStatus.SUCCESS.value,
                f"高风险预警已记录：reportId={report.id}，caseId={case.id if case else 'none'}，deliveryMode=log",
            )
        if mode != "smtp":
            return self._save_alert(
                report,
                recipient,
                ToolStatus.FAILED.value,
                f"高风险预警邮件未发送：未知投递模式 {self.settings.alert_email_delivery_mode}",
            )
        # 检查 SMTP 配置
        missing = self._missing_email_config()
        if missing:
            return self._save_alert(
                report,
                recipient,
                ToolStatus.FAILED.value,
                f"高风险预警邮件未发送：缺少配置 {', '.join(missing)}",
            )
        # 真正构造并发送邮件
        try:
            self._send_alert_email(report, case)
        # 发送邮件失败处理：SMTP 连接、登录或发送出错时：不让异常直接打断整个业务链路；
        # 保存一条失败的 AlertRecord；返回失败记录。
        except Exception as exc:
            return self._save_alert(
                report,
                recipient,
                ToolStatus.FAILED.value,
                f"高风险预警邮件发送失败：{type(exc).__name__}: {exc}",
            )
        return self._save_alert(report, recipient, ToolStatus.SUCCESS.value, f"高风险预警邮件已发送：reportId={report.id}")

    def _save_alert(self, report: PsychologicalReport, recipient: str, status: str, message: str) -> AlertRecord:
        record = AlertRecord(
            report_id=report.id,
            channel="email",
            recipient=recipient,
            status=status,
            message=message,
        )
        self.db.add(record)
        self.db.commit()
        return record

    def _add_case_note(self, case_id: int, actor: str, note: str) -> CaseNote:
        if not note:
            raise RuntimeError("case note cannot be empty")
        record = CaseNote(case_id=case_id, actor=actor, note=note)
        self.db.add(record)
        return record

    def _missing_email_config(self) -> list[str]:
        missing = []
        if not self.settings.smtp_host.strip():
            missing.append("SMTP_HOST")
        if not self._sender():
            missing.append("ALERT_EMAIL_FROM 或 SMTP_USERNAME")
        if not self._recipients():
            missing.append("ALERT_EMAIL_TO")
        return missing

    def _send_alert_email(self, report: PsychologicalReport, case: RiskCase | None = None) -> None:
        # 创建邮件对象
        message = EmailMessage()
        # 设置邮件头
        message["Subject"] = f"{self.settings.alert_email_subject_prefix} reportId={report.id}"
        message["From"] = self._sender()
        message["To"] = ", ".join(self._recipients())
        message.set_content(self._email_body(report, case))
        # 创建 TLS 安全上下文，使用系统默认的安全证书和 TLS 配置。
        context = ssl.create_default_context()
        # SSL 模式 SMTP_SSL + 465 端口
        if self.settings.smtp_use_ssl:
            with smtplib.SMTP_SSL(
                self.settings.smtp_host,
                self.settings.smtp_port,
                timeout=self.settings.smtp_timeout_seconds,
                context=context,
            ) as server:
                self._send_message(server, message)
            return
        # 普通 SMTP + STARTTLS 模式
        with smtplib.SMTP(self.settings.smtp_host, self.settings.smtp_port, timeout=self.settings.smtp_timeout_seconds) as server:
            # 连接普通 SMTP，向服务器声明客户端能力。
            server.ehlo()
            # 如果配置要求 TLS：升级为加密连接。SMTP + STARTTLS + 587 端口
            if self.settings.smtp_use_tls:
                server.starttls(context=context)
                server.ehlo()
            # 登录并发送
            self._send_message(server, message)

    def _send_message(self, server: smtplib.SMTP, message: EmailMessage) -> None:
        if self.settings.smtp_username:
            server.login(self.settings.smtp_username, self.settings.smtp_password)
        server.send_message(message)

# 构造邮件正文：作用是把报告、学生、个案和交接摘要组合成邮件正文。
    def _email_body(self, report: PsychologicalReport, case: RiskCase | None = None) -> str:
        user = self.db.get(UserAccount, report.user_id)
        username = user.username if user else f"userId={report.user_id}"
        display_name = user.display_name if user else ""
        # 获取交接摘要
        handoff = case.handoff_summary if case else MindBridgeSkillLibrary.counselor_handoff_summary(report, user)
        return "\n".join(
            [
                "CareTrace 检测到一条高风险心理预警，请尽快安排辅导员或管理员跟进。",
                "",
                f"个案ID：{case.id}" if case else "个案ID：未创建",
                f"报告ID：{report.id}",
                f"学生：{display_name} ({username})" if display_name else f"学生：{username}",
                f"风险等级：{report.risk_level}",
                f"情绪标签：{report.emotion}",
                f"置信度：{report.confidence}",
                f"摘要：{report.summary}",
                f"创建时间：{report.created_at.isoformat()}",
                "",
                "交接摘要：",
                handoff,
            ]
        )

    def _primary_owner(self) -> str:
        recipients = self._recipients()
        if recipients:
            return recipients[0]
        return "unassigned"

    def _sender(self) -> str:
        return self.settings.alert_email_from.strip() or self.settings.smtp_username.strip()

    def _recipients(self) -> list[str]:
        normalized = self.settings.alert_email_to.replace(";", ",")
        return [recipient.strip() for recipient in normalized.split(",") if recipient.strip()]
